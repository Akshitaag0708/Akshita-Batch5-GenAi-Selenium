import pytest
import openpyxl

# Create workbook
wb = openpyxl.Workbook()

sheetName = "sheet1"

# Check sheet exists or create
if sheetName in wb.sheetnames:
    ws = wb[sheetName]
else:
    ws = wb.create_sheet(sheetName)

ws["A1"] = "Username"
ws["B1"] = "Password"

wb.save("test.xlsx")

ws.append(["user1","1234"])
ws.append(["user2","1234"])
ws["a5"]="akshay"

# wb.save("test.xlsx")

# row wise
# for row in ws.iter_rows(values_only=True):      # to iterate in sheet
#     print(row)


# for row in ws.iter_rows(values_only=True):    # to iterate in sheet
#     for cell in row:
#         print(cell)


ws.delete_rows(3)
wb.save("test.xlsx")
for row in ws.iter_rows(values_only=True):      # to iterate in sheet
    print(row)

